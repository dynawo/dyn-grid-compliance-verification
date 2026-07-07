# Copyright (c) 2024-2026, RTE (https://www.rte-france.com)
# SPDX-License-Identifier: MPL-2.0
"""Fill a WECC template workbook with representative WECC values.

This is a helper to (re)build the committed example ``WECCSample.xlsx``: it
opens an existing WECC template (preserving its structure, formatting and the
optional ``Base unit`` / ``Comment`` columns), selects the blocks of a full
wind-turbine case and writes representative values for the selected variants
and the global parameters into a copy. The source template is never modified.

The values are invented (representative, not certified) and exist only to
exercise ``generate_par.py`` end to end.

Requires ``openpyxl`` (``uv pip install openpyxl``). The PAR generator itself
has no such dependency; this script is only for building example inputs.

Usage (regenerate the committed example next to this script)::

    python tools/dynawo_par/examples/fill_wecc_example.py \
        --src "Description_modèle_et_courbes (WECC) v0.xlsx"

    # ...or write somewhere else with --dst:
    python tools/dynawo_par/examples/fill_wecc_example.py --src template.xlsx --dst out.xlsx
    python tools/dynawo_par/generate_par.py --excel tools/dynawo_par/examples/WECCSample.xlsx
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path

import openpyxl

# --- Wind-turbine (WTG) case -------------------------------------------------
# A full wind turbine has BOTH the drive-train and the pitch / aerodynamic /
# torque control blocks. (A PV or BESS plant would leave those in 'Aucun', as
# it has no turbine; a BESS would also use the REEC_C battery variant.)

# Block selection written into the 'Type de bloc | Choix' table of 'Général'.
SELECTION = {
    "REPC": "REPC_A",
    "REEC": "REEC_A",
    "REGC": "REGC_A",
    "WTGT": "WTGT_B",
    "WTGP": "WTGP_A",
    "WTGA": "WTGA_A",
    "WTGQ": "WTGQ_A",
}

# Global parameters written into the 'Général' sheet (Valeur column).
# SnZone1 = 4 MVA per converter, 20 converters of 4.5 MVA, so the Zone 3
# contextual value is SnZone3 x Nombre de convertisseur = 90 MVA.
GLOBALS = {"SnZone1": "4", "Nombre de convertisseur": "20", "SnZone3": "4.5"}

# Representative WECC generic-model values, per selected variant
# ({parameter name: value}). Only these variants are populated; any other
# variant present in the template is left untouched.
VALUES: dict[str, dict[str, str]] = {
    "REPC_A": {
        "VCompFlag": "false", "RefFlag": "true", "FreqFlag": "true",
        "tFilterPC": "0.04", "Kp": "0.1", "Ki": "1.5", "tFt": "0", "tFv": "0.15",
        "VFrz": "0", "Kc": "0", "EMaxPu": "0.1", "EMinPu": "-0.1", "DbdPu": "0",
        "QMaxREPCPu": "0.436", "QMinREPCPu": "-0.436", "Kpg": "0.1", "Kig": "0.05",
        "tpREPC": "0.25", "FDbd1Pu": "0.000333", "FDbd2Pu": "0.000333",
        "FEMaxPu": "999", "FEMinPu": "-999", "PMaxREPCPu": "1.0", "PMinREPCPu": "0.0",
        "DDn": "20", "DUp": "0", "tLag": "0.1",
    },
    "REEC_A": {
        "PfFlag": "false", "VFlag": "true", "QFlag": "true", "PFlag": "false",
        "PQFlag": "false", "VDipPu": "0.9", "VUpPu": "1.1", "tRv": "0.01",
        "Dbd1Pu": "-0.05", "Dbd2Pu": "0.05", "Kqv": "2.0", "Iqh1Pu": "1.05",
        "Iql1Pu": "-1.05", "VRef0Pu": "1.0", "VRef1Pu": "0", "IqFrzPu": "0",
        "tHoldIq": "0", "tHoldIpMax": "0", "tpREEC": "0.05", "QMaxREECPu": "0.436",
        "QMinREECPu": "-0.436", "VMaxPu": "1.1", "VMinPu": "0.9", "Kqp": "1.0",
        "Kqi": "0.1", "Kvp": "1.0", "Kvi": "0.7", "tIq": "0.01", "DPMaxPu": "999",
        "DPMinPu": "-999", "PMaxREECPu": "1.0", "PMinREECPu": "0.0", "IMaxPu": "1.082",
        "tPord": "0.02",
        "VDLIq11": "0.0", "VDLIq12": "1.1", "VDLIq21": "0.2", "VDLIq22": "1.1",
        "VDLIq31": "0.5", "VDLIq32": "1.1", "VDLIq41": "1.0", "VDLIq42": "1.1",
        "VDLIp11": "0.0", "VDLIp12": "1.1", "VDLIp21": "0.2", "VDLIp22": "1.1",
        "VDLIp31": "0.5", "VDLIp32": "1.1", "VDLIp41": "1.0", "VDLIp42": "1.1",
    },
    "REGC_A": {
        "Iqrmax": "999", "Iqrmin": "-999", "Rrpwr": "10", "tFilterGC": "0.02",
        "tG": "0.02", "Lvplsw": "0", "brkpt": "0.9", "zerox": "0.4", "lvpl1": "1.22",
    },
    "WTGT_B": {  # Drive-Train (two-mass shaft)
        "Ht": "4.0", "Hg": "0.8", "Dshaft": "1.5", "Kshaft": "80", "tpWTGTb": "0.05",
    },
    "WTGP_A": {  # Pitch Control
        "Kiw": "25", "Kpw": "150", "Kic": "30", "Kpc": "3", "Kcc": "0",
        "tTheta": "0.3", "ThetaMax": "27", "ThetaMin": "0",
        "ThetaRMax": "10", "ThetaRMin": "-10",
    },
    "WTGA_A": {  # Aerodynamic
        "Ka": "0.007", "Theta0": "0",
    },
    "WTGQ_A": {  # Torque Control
        "TFlag": "true", "Kpp": "3", "Kip": "0.6", "tpWTGQa": "0.05",
        "tOmegaRef": "60", "TeMaxPu": "1.2", "TeMinPu": "0",
        "P1": "0.2", "Spd1": "0.58", "P2": "0.4", "Spd2": "0.72",
        "P3": "0.6", "Spd3": "0.86", "P4": "0.8", "Spd4": "1.0",
    },
}


def _norm(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _fill_selection(ws) -> None:
    """Set the chosen variant of each block in the 'Type de bloc | Choix' table."""
    for r in range(1, ws.max_row + 1):
        labels = [_norm(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        if "type de bloc" in labels and "choix" in labels:
            block_col = labels.index("type de bloc") + 1
            choice_col = labels.index("choix") + 1
            for rr in range(r + 1, ws.max_row + 1):
                block = _norm(ws.cell(rr, block_col).value)
                if not block:
                    break
                if block in SELECTION:
                    ws.cell(rr, choice_col).value = SELECTION[block]
            return


def _fill_globals(ws) -> None:
    """Set the global parameters in the 'Grandeur | … | Valeur' table."""
    for r in range(1, ws.max_row + 1):
        labels = [_norm(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        if "grandeur" in labels and "valeur" in labels:
            name_col = labels.index("grandeur") + 1
            value_col = labels.index("valeur") + 1
            for rr in range(r + 1, ws.max_row + 1):
                name = _norm(ws.cell(rr, name_col).value)
                if not name:
                    break
                if name in GLOBALS:
                    ws.cell(rr, value_col).value = GLOBALS[name]
            return


def _fill_params(ws) -> None:
    """Fill the Value column of each selected variant by parameter name."""
    header = next(
        (
            r
            for r in range(1, ws.max_row + 1)
            if any(
                _norm(ws.cell(r, c).value).lower() == "parameter"
                for c in range(1, ws.max_column + 1)
            )
        ),
        None,
    )
    if header is None:
        return
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(header, c).value).lower() != "parameter":
            continue
        variant = _norm(ws.cell(header - 1, c).value)
        table = VALUES.get(variant)
        if not table:
            continue
        value_col = c + 2
        for r in range(header + 1, ws.max_row + 1):
            name = _norm(ws.cell(r, c).value)
            if name in table:
                ws.cell(r, value_col).value = table[name]


def fill(src: Path, dst: Path) -> None:
    shutil.copy(src, dst)
    workbook = openpyxl.load_workbook(dst)
    for ws in workbook.worksheets:
        if ws.title.strip().lower().startswith(("general", "général")):
            _fill_selection(ws)
            _fill_globals(ws)
        else:
            _fill_params(ws)
    workbook.save(dst)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--src", required=True, type=Path, help="WECC template .xlsx.")
    parser.add_argument(
        "--dst",
        type=Path,
        default=None,
        help="Output .xlsx (default: WECCSample.xlsx next to this script).",
    )
    args = parser.parse_args()
    if not args.src.is_file():
        parser.error(f"Template not found: {args.src}")
    dst = args.dst or Path(__file__).resolve().parent / "WECCSample.xlsx"
    fill(args.src, dst)
    print(f"Wrote {dst}")


if __name__ == "__main__":
    main()
