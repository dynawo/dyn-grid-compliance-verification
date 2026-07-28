#!/usr/bin/env python3
# Copyright (c) 2024-2026, RTE (https://www.rte-france.com)
# SPDX-License-Identifier: MPL-2.0
"""Build the AIA-authored test fixture ``WECCSample_full.xlsx`` for the input generator.

This is a **testing artifact with representative contents** — columns, parameter names *and*
values — used only to exercise the tool end to end. It is independent of the production template's
responsibility split (AIA=columns / RTE=names / end user=values).

The case mirrors the committed example ``examples/Model/Photovoltaics/WECCCurrentSource``: a 90 MVA
PV ``S+Aux`` plant (``PhotovoltaicsWeccCurrentSource``), so the generated input tree resembles that
example as closely as the Excel model allows. The control-block parameters and their values are
taken from that example's ``Producer.par``; the sheet layout follows the original WECC template
(a table-name row, a variant row, a ``Parameter | Type | Value`` header with optional ``Base unit``
/ ``Comment`` columns). The electrical values are chosen so the computed per-unit impedances match
the example (e.g. the step-up ``XPu ~= 0.027`` on ``SnRef = 100 MVA``).

Run with a Python that has ``openpyxl``; the committed ``.xlsx`` is regenerated deterministically.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

OUT = Path(__file__).resolve().parent / "WECCSample_full.xlsx"

# Full variant -> (Zone3 lib/prefix, Zone1 lib/prefix) map (10 WECC producer models).
MODEL_MAP_ROWS = [
    ("REGC_A|REEC_B|Aucun|Aucun|Aucun|Aucun", "PhotovoltaicsWeccCurrentSource", "photovoltaics_",
     "PhotovoltaicsWeccCurrentSourceNoPlantControl", "photovoltaics_"),
    ("REGC_B|REEC_A|Aucun|Aucun|Aucun|Aucun", "PhotovoltaicsWeccVoltageSource1", "photovoltaics_",
     "PhotovoltaicsWeccVoltageSource1NoPlantControl", "photovoltaics_"),
    ("REGC_B|REEC_B|Aucun|Aucun|Aucun|Aucun", "PhotovoltaicsWeccVoltageSource2", "photovoltaics_",
     "PhotovoltaicsWeccVoltageSource2NoPlantControl", "photovoltaics_"),
    ("REGC_C|REEC_A|Aucun|Aucun|Aucun|Aucun", "PhotovoltaicsWeccVoltageSource3", "photovoltaics_",
     "PhotovoltaicsWeccVoltageSource3NoPlantControl", "photovoltaics_"),
    ("REGC_C|REEC_B|Aucun|Aucun|Aucun|Aucun", "PhotovoltaicsWeccVoltageSource4", "photovoltaics_",
     "PhotovoltaicsWeccVoltageSource4NoPlantControl", "photovoltaics_"),
    ("REGC_A|REEC_C|Aucun|Aucun|Aucun|Aucun", "BESSWeccCurrentSource", "BESS_",
     "BESSWeccCurrentSourceNoPlantControl", "BESS_"),
    ("REGC_A|REEC_A|WTGT_A|WTGP_A|WTGA_A|WTGQ_A", "WTG3WeccCurrentSource1", "WTG3_",
     "WeccWT3CurrentSource1", "WT3_"),
    ("REGC_A|REEC_A|WTGT_A|WTGP_B|WTGA_A|WTGQ_A", "WTG3WeccCurrentSource2", "WTG3_",
     "WeccWT3CurrentSource2", "WT3_"),
    ("REGC_A|REEC_A|WTGT_B|Aucun|Aucun|Aucun", "WTG4AWeccCurrentSource", "WTG4A_",
     "WT4AWeccCurrentSource", "WT4A_"),
    ("REGC_A|REEC_A|Aucun|Aucun|Aucun|Aucun", "WTG4BWeccCurrentSource", "WTG4B_",
     "WT4BWeccCurrentSource", "WT4B_"),
]

# Zone1a: name in col A, value in col C (per design 4.3). One 90 MVA PV unit, ConverterLVControl
# True (so Z_cc_TG is the explicit StepUp_Xfmr). Z_cc_TG/R_cc give XPu ~= 0.027, RPu ~= 0.0003 on
# SnRef = 100 MVA (base s_nom = SnZone1 = 90), matching the example's step-up transformer.
ZONE1 = [
    ("SnZone1", 90, "MVA"), ("N_Zone1", 1, "-"), ("ConverterLVControl", "True", "-"),
    ("Un2", 0.69, "kV"), ("Un1", 33, "kV"),
    # Internal LV transformer (LvTr): neutral placeholder (the end user sets the real impedance).
    ("Z_cc_LvTr", 0.0001, "pu"), ("R_cc_LvTr / X_cc_LvTr", 0, "-"),
    # Generator step-up transformer (StepUp_Xfmr, fixed ratio).
    ("r_TG", 1, "pu"), ("Z_cc_TG", 0.02445, "pu"), ("R_cc_TG / X_cc_TG", 0.01115, "-"),
    ("Pmax_injection_z1", 90, "MW"), ("Pmax_soutirage_z1", 0, "MW"),
    ("Qmax_z1", 30, "MVAr"), ("Qmin_z1", -30, "MVAr"), ("P_share", 1, "-"), ("Q_share", 1, "-"),
]
# Zone3: name in col B, value in col D (per design 4.3). 90 MVA plant, S+Aux; the near-ideal
# auxiliary transformer/load reproduce the example (XPu ~= 1e-4, load 0.01/0.005 pu on SnRef=100).
ZONE3 = [
    ("Paramètres généraux", "SnZone3", 90, "MVA"), ("", "Topologie", "S+Aux", "-"),
    ("", "Un_PDR", 225, "kV"), ("", "Pmax_PDR", 90, "MW"), ("", "Qmax_PDR", 30, "MVAr"),
    ("", "Qmin_PDR", -30, "MVAr"),
    ("Transformateur principal", "Z_cc_TP", 0.12, "pu"), ("", "R_cc_TP / X_cc_TP", 0.01, "-"),
    ("", "N_prises", 20, "-"), ("", "r_max", 1.1, "pu"), ("", "r_min", 0.9, "pu"),
    ("Charge auxiliaire", "Un_A", 0.69, "kV"), ("", "Sn_A", 1, "MVA"), ("", "r_TA", 1, "pu"),
    ("", "Z_cc_TA", 0.000001, "pu"), ("", "R_cc_TA / X_cc_TA", 0.1, "-"), ("", "P_A", 1, "MW"),
    ("", "Q_A", 0.5, "MVAr"), ("", "alpha", 1.0, "-"), ("", "beta", 1.0, "-"),
]

# Control blocks: {sheet: (variant, table name, [(name, type, value, base_unit, comment)])}.
# Values are the example's PhotovoltaicsWeccCurrentSource Producer.par (bare names; the model prefix
# is prepended by the tool). REPC is the plant controller (dropped from the Zone1 turbine output).
CONTROL = {
    "REEC": ("REEC_B", "Electrical Control", [
        ("PfFlag", "boolean", "false", "", "power-factor control flag"),
        ("PQFlag", "boolean", "false", "", "P/Q priority flag"),
        ("QFlag", "boolean", "true", "", "reactive control flag"),
        ("VFlag", "boolean", "true", "", "voltage control flag"),
        ("Dbd1Pu", "double", "-0.1", "Un", ""),
        ("Dbd2Pu", "double", "0.1", "Un", ""),
        ("DPMaxPu", "double", "999", "SNom", ""),
        ("DPMinPu", "double", "-999", "SNom", ""),
        ("IMaxPu", "double", "1.05", "SNom", ""),
        ("Iqh1Pu", "double", "2.0", "SNom", ""),
        ("Iql1Pu", "double", "-2.0", "SNom", ""),
        ("Kqi", "double", "0.5", "", ""),
        ("Kqp", "double", "1", "", ""),
        ("Kqv", "double", "2", "", ""),
        ("Kvi", "double", "1", "", ""),
        ("Kvp", "double", "1", "", ""),
        ("PMaxREECPu", "double", "1.0", "SNom", ""),
        ("PMinREECPu", "double", "0.0", "SNom", ""),
        ("QMaxREECPu", "double", "0.5", "SNom", ""),
        ("QMinREECPu", "double", "-0.5", "SNom", ""),
        ("PMaxREPCPu", "double", "1.0", "SNom", ""),
        ("PMinREPCPu", "double", "0.0", "SNom", ""),
        ("QMaxREPCPu", "double", "0.5", "SNom", ""),
        ("QMinREPCPu", "double", "-0.5", "SNom", ""),
        ("tIq", "double", "0.02", "s", ""),
        ("tRv", "double", "0.02", "s", ""),
        ("tpREEC", "double", "0.04", "s", ""),
        ("tpREPC", "double", "0.04", "s", ""),
        ("tPord", "double", "0.02", "s", ""),
        ("VDipPu", "double", "0.9", "Un", ""),
        ("VMaxPu", "double", "1.1", "Un", ""),
        ("VMinPu", "double", "0.9", "Un", ""),
        ("VRef0Pu", "double", "1", "Un", ""),
        ("VUpPu", "double", "1.35", "Un", ""),
        ("VRef1Pu", "double", "0", "Un", "REEC_B-specific reference"),
    ]),
    "REGC": ("REGC_A", "Generator Converter", [
        ("IqrMaxPu", "double", "20", "SNom", ""),
        ("IqrMinPu", "double", "-20", "SNom", ""),
        ("RrpwrPu", "double", "10", "SNom", ""),
        ("tFilterGC", "double", "0.02", "s", ""),
        ("tG", "double", "0.02", "s", ""),
        ("brkpt", "double", "0.05", "Un", ""),
        ("lvpl1", "double", "1.22", "SNom", ""),
        ("Lvplsw", "boolean", "false", "", "low-voltage power-logic switch"),
        ("zerox", "double", "0.1", "Un", ""),
        ("KiPLL", "double", "20", "", "PLL integral gain"),
        ("KpPLL", "double", "3", "", "PLL proportional gain"),
        ("OmegaMaxPu", "double", "1.5", "OmegaNom", ""),
        ("OmegaMinPu", "double", "0.5", "OmegaNom", ""),
    ]),
    "REPC": ("REPC_A", "Plant Control", [
        ("FreqFlag", "boolean", "true", "", "frequency control flag"),
        ("RefFlag", "boolean", "true", "", "voltage/Q reference flag"),
        ("VCompFlag", "boolean", "false", "", "line-drop compensation flag"),
        ("DbdPu", "double", "0.0001", "Un", ""),
        ("DDn", "double", "100", "", ""),
        ("DUp", "double", "100", "", ""),
        ("EMaxPu", "double", "0.5", "SNom", ""),
        ("EMinPu", "double", "-0.5", "SNom", ""),
        ("FDbd1Pu", "double", "0.001", "fNom", ""),
        ("FDbd2Pu", "double", "0.001", "fNom", ""),
        ("FEMaxPu", "double", "999", "fNom", ""),
        ("FEMinPu", "double", "-999", "fNom", ""),
        ("Kc", "double", "0.3", "", ""),
        ("Ki", "double", "1.5", "", ""),
        ("Kig", "double", "2.36", "", ""),
        ("Kp", "double", "0.1", "", ""),
        ("Kpg", "double", "0.05", "", ""),
        ("tFilterPC", "double", "0.04", "s", ""),
        ("tFt", "double", "1e-5", "s", ""),
        ("tFv", "double", "0.1", "s", ""),
        ("tLag", "double", "0.1", "s", ""),
        ("VFrz", "double", "0", "Un", ""),
    ]),
}


def build() -> Path:
    wb = Workbook()

    gen = wb.active
    gen.title = "Général"
    gen.append(["Type de bloc", "Choix"])
    for block, choice in [("REPC", "REPC_A"), ("REEC", "REEC_B"), ("REGC", "REGC_A"),
                          ("WTGT", "Aucun"), ("WTGP", "Aucun"), ("WTGA", "Aucun"),
                          ("WTGQ", "Aucun")]:
        gen.append([block, choice])

    mm = wb.create_sheet("Model Map")
    mm.append(["Key", "Zone3_lib", "Zone3_prefix", "Zone1_lib", "Zone1_prefix"])
    for row in MODEL_MAP_ROWS:
        mm.append(list(row))

    z1 = wb.create_sheet("Zone1a")
    z1.append(["Le schéma de base pour la zone 1 est le suivant :"])
    z1.append(["Paramètres", "Descriptions", "Valeurs", "Unités", "Commentaires"])
    for name, value, unit in ZONE1:
        z1.append([name, "", value, unit, ""])

    z3 = wb.create_sheet("Zone3")
    z3.append(["- Topologies ..."])
    z3.append(["Catégorie", "Paramètres", "Descriptions", "Valeurs", "Unités", "Commentaires"])
    for category, name, value, unit in ZONE3:
        z3.append([category, name, "", value, unit, ""])

    for sheet, (variant, table, params) in CONTROL.items():
        ws = wb.create_sheet(sheet)
        ws.append([table])  # table name (row N-2)
        ws.append([variant])  # variant name, above the Parameter column (row N-1)
        ws.append(["Parameter", "Type", "Value", "Base unit", "Comment"])
        for name, typ, val, base, comment in params:
            ws.append([name, typ, val, base, comment])

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print("wrote", build())
